"""
generate_attendance_excel.py
-----------------------------
Called after every face-recognition session.

Steps:
1. Pull every student from the MySQL `student` table.
2. Read today's Attend.csv to find who was marked Present.
3. Anyone NOT in the Present list is written as Absent.
4. Save a dated Excel file: Attendance_YYYY-MM-DD.xlsx
"""

import csv
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import mysql.connector


# ── DB connection settings (same as the rest of the project) ──────────────────
DB_CONFIG = dict(
    host="localhost",
    user="root",
    password="admin",
    database="face_recognizer",
)

ATTEND_CSV  = "Attend.csv"          # path to the running attendance CSV
OUTPUT_DIR  = "attendance_reports"  # folder where Excel files are saved


def _fetch_all_students():
    """Return list of dicts for every student in the DB."""
    conn = mysql.connector.connect(**DB_CONFIG)
    cur  = conn.cursor(dictionary=True)
    cur.execute(
        "SELECT studentID, student_name, Roll_no, dep, section, semester "
        "FROM student ORDER BY Roll_no"
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def _fetch_present_today(today_str: str) -> set:
    """
    Return a set of studentIDs that are marked Present in Attend.csv
    for the given date (format DD/MM/YYYY).
    """
    present = set()
    if not os.path.exists(ATTEND_CSV):
        return present

    with open(ATTEND_CSV, newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            # CSV columns: studentID, Roll_no, student_name, dep, time, date, status
            if len(row) >= 7:
                sid   = str(row[0]).strip()
                date  = str(row[5]).strip()
                status = str(row[6]).strip().lower()
                if date == today_str and status == "present":
                    present.add(sid)
    return present


def generate_excel(output_path: str = None) -> str:
    """
    Build the final attendance Excel workbook.
    Returns the path of the saved file.
    """
    today     = datetime.now()
    today_str = today.strftime("%d/%m/%Y")      # matches Attend.csv format
    date_tag  = today.strftime("%Y-%m-%d")

    students = _fetch_all_students()
    present  = _fetch_present_today(today_str)

    # ── Merge: every student gets Present or Absent ───────────────────────────
    records = []
    for s in students:
        sid    = str(s["studentID"])
        status = "Present" if sid in present else "Absent"
        records.append({
            "studentID"   : sid,
            "Roll_no"     : s["Roll_no"],
            "student_name": s["student_name"],
            "dep"         : s["dep"],
            "section"     : s["section"],
            "semester"    : s["semester"],
            "date"        : today_str,
            "status"      : status,
        })

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Colour palette
    HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
    PRESENT_FILL  = PatternFill("solid", fgColor="C6EFCE")
    ABSENT_FILL   = PatternFill("solid", fgColor="FFC7CE")
    ALT_ROW_FILL  = PatternFill("solid", fgColor="EBF3FB")
    WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")

    HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    BODY_FONT     = Font(name="Calibri", size=10)
    TITLE_FONT    = Font(name="Calibri", bold=True, size=14, color="1F4E79")

    thin = Side(style="thin", color="B8CCE4")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center")

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value     = f"FACE RECOGNITION ATTENDANCE SYSTEM — {today_str}"
    title_cell.font      = TITLE_FONT
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # ── Summary row ───────────────────────────────────────────────────────────
    total   = len(records)
    n_pres  = sum(1 for r in records if r["status"] == "Present")
    n_abs   = total - n_pres

    ws.merge_cells("A2:H2")
    summary_cell = ws["A2"]
    summary_cell.value = (
        f"Total: {total}   |   Present: {n_pres}   |   Absent: {n_abs}   "
        f"|   Date: {today_str}"
    )
    summary_cell.font      = Font(name="Calibri", bold=True, size=10, color="595959")
    summary_cell.alignment = CENTER
    ws.row_dimensions[2].height = 18

    # ── Column headers (row 3) ─────────────────────────────────────────────────
    headers = [
        "S.No", "Student ID", "Roll No", "Student Name",
        "Department", "Section", "Semester", "Attendance Status"
    ]
    col_widths = [6, 12, 10, 22, 16, 10, 12, 18]

    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[3].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, rec in enumerate(records, start=1):
        excel_row = row_idx + 3          # offset for title + summary + header
        is_present = rec["status"] == "Present"

        row_fill = PRESENT_FILL if is_present else ABSENT_FILL
        # Use alternating white/light-blue only for present rows when not absent
        if is_present and row_idx % 2 == 0:
            row_fill = ALT_ROW_FILL if is_present else ABSENT_FILL

        values = [
            row_idx,
            rec["studentID"],
            rec["Roll_no"],
            rec["student_name"],
            rec["dep"],
            rec["section"],
            rec["semester"],
            rec["status"],
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font      = BODY_FONT
            cell.border    = BORDER
            cell.alignment = CENTER

            # Status column gets strong colour; other columns get subtle tint
            if col_idx == 8:
                cell.fill = PRESENT_FILL if is_present else ABSENT_FILL
                cell.font = Font(
                    name="Calibri", bold=True, size=10,
                    color="375623" if is_present else "9C0006"
                )
            elif is_present:
                cell.fill = WHITE_FILL if row_idx % 2 else ALT_ROW_FILL
            else:
                cell.fill = PatternFill("solid", fgColor="FFE0E0")

        ws.row_dimensions[excel_row].height = 16

    # ── Freeze panes below header ─────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Auto-filter on header row ─────────────────────────────────────────────
    ws.auto_filter.ref = f"A3:H{3 + len(records)}"

    # ── Save ──────────────────────────────────────────────────────────────────
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if output_path is None:
        output_path = os.path.join(OUTPUT_DIR, f"Attendance_{date_tag}.xlsx")

    wb.save(output_path)
    print(f"[AttendanceExcel] Saved → {output_path}")
    return output_path


# Standalone test
if __name__ == "__main__":
    path = generate_excel()
    print(f"Done: {path}")
