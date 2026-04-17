"""
export_students_excel.py
------------------------
Exports the full student master list from MySQL to
  student_reports/All_Students.xlsx

Called automatically from student.py on Save / Update / Delete.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import mysql.connector
from datetime import datetime

DB_CONFIG = dict(
    host="localhost",
    user="root",
    password="admin",
    database="face_recognizer",
)

OUTPUT_DIR = "student_reports"


def export_students(output_path: str = None) -> str:
    conn = mysql.connector.connect(**DB_CONFIG)
    cur  = conn.cursor()
    cur.execute("SELECT * FROM student ORDER BY Roll_no")
    rows    = cur.fetchall()
    columns = [desc[0] for desc in cur.description]
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Students"

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    BODY_FONT   = Font(name="Calibri", size=10)
    ALT_FILL    = PatternFill("solid", fgColor="EBF3FB")
    WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
    thin        = Side(style="thin", color="B8CCE4")
    BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER      = Alignment(horizontal="center", vertical="center")

    # Title
    ws.merge_cells(f"A1:{chr(64+len(columns))}1")
    t = ws["A1"]
    t.value     = f"STUDENT MASTER LIST — Generated {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    t.font      = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26

    # Headers
    for ci, col in enumerate(columns, 1):
        cell = ws.cell(row=2, column=ci, value=col.replace("_", " ").title())
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[cell.column_letter].width = max(14, len(col) + 4)
    ws.row_dimensions[2].height = 18

    # Data
    for ri, row in enumerate(rows, 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri + 2, column=ci, value=val)
            cell.font      = BODY_FONT
            cell.alignment = CENTER
            cell.border    = BORDER
            cell.fill      = WHITE_FILL if ri % 2 else ALT_FILL
        ws.row_dimensions[ri + 2].height = 15

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{chr(64+len(columns))}{2+len(rows)}"

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if output_path is None:
        output_path = os.path.join(OUTPUT_DIR, "All_Students.xlsx")
    wb.save(output_path)
    print(f"[StudentsExcel] Saved → {output_path}")
    return output_path


if __name__ == "__main__":
    print(export_students())
